import openpyxl
import os

def compare_xlsx_files(file1_path, file2_path, output_path=None):
    try:
        # Load workbooks from both files
        wb1 = openpyxl.load_workbook(file1_path)
        wb2 = openpyxl.load_workbook(file2_path)

        # Create an output workbook or load the existing one
        if output_path and os.path.exists(output_path):
            output_wb = openpyxl.load_workbook(output_path)
        else:
            output_wb = openpyxl.Workbook()

        # Clear existing sheets in the output workbook
        for sheet_name in output_wb.sheetnames:
            output_wb.remove(output_wb[sheet_name])

        # Get sheet names from both files
        sheet_names1 = wb1.sheetnames
        sheet_names2 = wb2.sheetnames

        # Iterate through each sheet in both files
        for sheet_name in sheet_names1:
            # Get the sheets from both files, if present
            sheet1 = wb1[sheet_name] if sheet_name in sheet_names1 else None
            sheet2 = wb2[sheet_name] if sheet_name in sheet_names2 else None

            # Create a new sheet in the output workbook
            output_sheet = output_wb.create_sheet(title=sheet_name)

            # Write the header row to indicate the source file
            output_sheet.append(["Data", "Source"])

            # Get data from both sheets
            if sheet1:
                data1 = sheet1.iter_rows(values_only=True)
                for row in data1:
                    output_sheet.append([cell_value for cell_value in row] + ["file1"])

            if sheet2:
                data2 = sheet2.iter_rows(values_only=True)
                for row in data2:
                    output_sheet.append([cell_value for cell_value in row] + ["file2"])

        # Save the output workbook
        if output_path is None:
            # If no output path provided, save in the same directory as the script with a default name
            output_path = os.path.join(os.path.dirname(__file__), "output.xlsx")

        output_wb.save(output_path)
        print(f"Comparison completed. Results saved to {output_path}")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    file1_path = input("Enter the path of file1.xlsx: ")
    file2_path = input("Enter the path of file2.xlsx: ")

    if not (os.path.isfile(file1_path) and os.path.isfile(file2_path)):
        print("One or both input files do not exist. Please provide valid file paths.")
    else:
        output_path = input("Enter the path for the output file (press Enter for default location): ")
        compare_xlsx_files(file1_path, file2_path, output_path)
