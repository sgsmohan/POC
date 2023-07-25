import openpyxl
import os

def compare_xlsx_files(file1_path, file2_path, output_path=None):
    try:
        # Load workbooks from both files
        wb1 = openpyxl.load_workbook(file1_path)
        wb2 = openpyxl.load_workbook(file2_path)

        # Create an output workbook
        output_wb = openpyxl.Workbook()

        # Get sheet names from both files
        sheet_names1 = wb1.sheetnames
        sheet_names2 = wb2.sheetnames

        # Ensure that the output workbook has sheets corresponding to all sheets in both files
        for sheet_name in sheet_names1:
            if sheet_name not in output_wb.sheetnames:
                output_wb.create_sheet(title=sheet_name)

        for sheet_name in sheet_names2:
            if sheet_name not in output_wb.sheetnames:
                output_wb.create_sheet(title=sheet_name)

        # Iterate through each sheet in both files
        for sheet_name in output_wb.sheetnames:
            # Get the sheets from both files, if present
            sheet1 = wb1[sheet_name] if sheet_name in sheet_names1 else None
            sheet2 = wb2[sheet_name] if sheet_name in sheet_names2 else None

            # Create a dictionary to store data from both sheets
            data_dict = {}

            # Get data from both sheets
            if sheet1:
                data1 = sheet1.iter_rows(values_only=True)
                data_dict["file1"] = set(data1)
            else:
                data_dict["file1"] = set()

            if sheet2:
                data2 = sheet2.iter_rows(values_only=True)
                data_dict["file2"] = set(data2)
            else:
                data_dict["file2"] = set()

            # Find the rows present in file1 but missing in file2
            missing_in_file2 = data_dict["file1"] - data_dict["file2"]

            # Find the rows present in file2 but missing in file1
            missing_in_file1 = data_dict["file2"] - data_dict["file1"]

            # Write the differences to the output sheets
            output_sheet1 = output_wb[sheet_name]
            output_sheet1.append(["Data missing in file2 from file1:"])
            for row in missing_in_file2:
                output_sheet1.append(row)
            
            output_sheet2 = output_wb.create_sheet(title=f"{sheet_name}_missing_in_file1")
            output_sheet2.append(["Data missing in file1 from file2:"])
            for row in missing_in_file1:
                output_sheet2.append(row)

        # Save the output workbook
        if output_path is None:
            # If no output path provided, save in the same directory as the script with a default name
            output_path = os.path.join(os.path.dirname(__file__), "output.xlsx")

        output_wb.save(output_path)
        print(f"Comparison completed. Results saved to {output_path}")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    file1_path = input("Enter the path of file1.xlsx: ")
    file2_path = input("Enter the path of file2.xlsx: ")

    if not (os.path.isfile(file1_path) and os.path.isfile(file2_path)):
        print("One or both input files do not exist. Please provide valid file paths.")
    else:
        output_path = input("Enter the path for the output file (press Enter for default location): ")
        compare_xlsx_files(file1_path, file2_path, output_path)
